"""
FlexoBrain Knowledge Loader - Multi-source Data Import

This module handles loading knowledge from various sources:
- PDF documents (technical data sheets, spec sheets)
- CSV/Excel files (bulk plate data, specifications)
- Text files and documents
- Manual knowledge entries
- OpenAI file_search uploads for RAG

Usage:
    - Upload files via API endpoints
    - Bulk import from CSV
    - Seed with industry knowledge
"""

import os
import io
import json
import hashlib
import csv
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks
from pydantic import BaseModel
import asyncpg

router = APIRouter(prefix="/api/knowledge/load", tags=["Knowledge Loader"])

# ============================================================================
# CONFIGURATION
# ============================================================================

UPLOAD_DIR = "/tmp/flexobrain_uploads"
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB max

# Supported file types
SUPPORTED_EXTENSIONS = {
    'pdf': 'application/pdf',
    'txt': 'text/plain',
    'md': 'text/markdown',
    'csv': 'text/csv',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'json': 'application/json'
}

# ============================================================================
# DATABASE CONNECTION
# ============================================================================

_db_pool = None

async def get_db_pool():
    global _db_pool
    if _db_pool is None:
        database_url = os.getenv("DATABASE_URL")
        if database_url:
            _db_pool = await asyncpg.create_pool(database_url, min_size=2, max_size=10)
    return _db_pool


# ============================================================================
# OPENAI CLIENT
# ============================================================================

_openai_client = None

def get_openai_client():
    global _openai_client
    if _openai_client is None:
        from openai import AsyncOpenAI
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not set")
        _openai_client = AsyncOpenAI(api_key=api_key)
    return _openai_client


# ============================================================================
# PYDANTIC MODELS
# ============================================================================

class ManualKnowledgeEntry(BaseModel):
    title: str
    content: str
    category: str  # plates, processing, equipment, troubleshooting, best_practices
    subcategory: Optional[str] = None
    tags: Optional[List[str]] = None
    supplier_name: Optional[str] = None

class BulkPlateData(BaseModel):
    plates: List[Dict[str, Any]]

class SeedKnowledgeRequest(BaseModel):
    include_core_knowledge: bool = True
    include_supplier_info: bool = True
    include_troubleshooting: bool = True
    include_best_practices: bool = True

class ImportCSVRequest(BaseModel):
    category: str = "plates"
    has_header: bool = True


# ============================================================================
# PDF PROCESSING
# ============================================================================

async def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF using PyPDF2"""
    try:
        import pypdf

        pdf_reader = pypdf.PdfReader(io.BytesIO(file_content))
        text_parts = []

        for page in pdf_reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)

        return '\n\n'.join(text_parts)
    except ImportError:
        # Fallback: try pdfplumber
        try:
            import pdfplumber

            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                text_parts = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                return '\n\n'.join(text_parts)
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="PDF processing libraries not installed. Install pypdf or pdfplumber."
            )


# ============================================================================
# EXCEL PROCESSING
# ============================================================================

async def parse_excel_file(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse Excel file to list of dicts"""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            if any(row_dict.values()):  # Skip empty rows
                rows.append(row_dict)

        return rows
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel processing library not installed. Install openpyxl."
        )


# ============================================================================
# TEXT CHUNKING & EMBEDDING (shared with knowledge_scraper)
# ============================================================================

def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> List[Dict[str, Any]]:
    """Split text into chunks for embedding"""
    try:
        import tiktoken
        tokenizer = tiktoken.encoding_for_model("gpt-4")
    except:
        # Fallback to word-based chunking
        words = text.split()
        chunks = []
        for i in range(0, len(words), chunk_size - overlap):
            chunk_words = words[i:i + chunk_size]
            chunks.append({
                "text": ' '.join(chunk_words),
                "tokens": len(chunk_words),
                "index": len(chunks)
            })
        return chunks

    import re
    sentences = re.split(r'(?<=[.!?])\s+', text)

    chunks = []
    current_chunk = []
    current_tokens = 0

    for sentence in sentences:
        sentence_tokens = len(tokenizer.encode(sentence))

        if current_tokens + sentence_tokens > chunk_size:
            if current_chunk:
                chunk_text = ' '.join(current_chunk)
                chunks.append({
                    "text": chunk_text,
                    "tokens": current_tokens,
                    "index": len(chunks)
                })
            current_chunk = [sentence]
            current_tokens = sentence_tokens
        else:
            current_chunk.append(sentence)
            current_tokens += sentence_tokens

    if current_chunk:
        chunk_text = ' '.join(current_chunk)
        chunks.append({
            "text": chunk_text,
            "tokens": current_tokens,
            "index": len(chunks)
        })

    return chunks


async def generate_embeddings(texts: List[str]) -> List[List[float]]:
    """Generate embeddings for texts"""
    client = get_openai_client()

    batch_size = 100
    all_embeddings = []

    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        response = await client.embeddings.create(
            model="text-embedding-3-small",
            input=batch,
            dimensions=1536
        )
        batch_embeddings = [item.embedding for item in response.data]
        all_embeddings.extend(batch_embeddings)

    return all_embeddings


def content_hash(content: str) -> str:
    """Generate hash for deduplication"""
    return hashlib.sha256(content.encode('utf-8')).hexdigest()


async def check_has_vector_column(conn) -> bool:
    """Check if the embedding column exists in knowledge_chunks table"""
    try:
        result = await conn.fetchval("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'knowledge_chunks' AND column_name = 'embedding'
            )
        """)
        return result
    except:
        return False


async def store_knowledge(
    pool: asyncpg.Pool,
    title: str,
    content: str,
    source_type: str,
    source_name: str,
    category: str,
    subcategory: str = None,
    tags: List[str] = None,
    source_url: str = None,
    supplier_id: str = None
) -> Tuple[str, bool]:
    """Store document with chunks and embeddings (if pgvector available)"""
    hash_val = content_hash(content)

    async with pool.acquire() as conn:
        # Check duplicate
        existing = await conn.fetchrow(
            "SELECT id FROM knowledge_documents WHERE content_hash = $1",
            hash_val
        )

        if existing:
            return str(existing['id']), False

        # Insert document
        doc_id = await conn.fetchval("""
            INSERT INTO knowledge_documents
            (source_url, source_type, source_name, title, content, content_hash,
             category, subcategory, tags, supplier_id, word_count)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)
            RETURNING id
        """, source_url, source_type, source_name, title, content, hash_val,
            category, subcategory, tags, supplier_id, len(content.split()))

        # Chunk the content
        chunks = chunk_text(content)

        if chunks:
            # Check if we have vector support
            has_vector = await check_has_vector_column(conn)

            if has_vector:
                # Generate embeddings and store with vectors
                try:
                    chunk_texts = [c['text'] for c in chunks]
                    embeddings = await generate_embeddings(chunk_texts)

                    for chunk, embedding in zip(chunks, embeddings):
                        await conn.execute("""
                            INSERT INTO knowledge_chunks
                            (document_id, chunk_index, chunk_text, chunk_tokens, embedding)
                            VALUES ($1, $2, $3, $4, $5)
                        """, doc_id, chunk['index'], chunk['text'], chunk['tokens'],
                            json.dumps(embedding))
                except Exception as e:
                    print(f"Error storing embeddings, falling back to text-only: {e}")
                    # Fallback: store without embeddings
                    for chunk in chunks:
                        await conn.execute("""
                            INSERT INTO knowledge_chunks
                            (document_id, chunk_index, chunk_text, chunk_tokens)
                            VALUES ($1, $2, $3, $4)
                        """, doc_id, chunk['index'], chunk['text'], chunk['tokens'])
            else:
                # No vector support - store chunks without embeddings
                for chunk in chunks:
                    await conn.execute("""
                        INSERT INTO knowledge_chunks
                        (document_id, chunk_index, chunk_text, chunk_tokens)
                        VALUES ($1, $2, $3, $4)
                    """, doc_id, chunk['index'], chunk['text'], chunk['tokens'])

        return str(doc_id), True


# ============================================================================
# API ENDPOINTS
# ============================================================================

@router.post("/manual")
async def add_manual_knowledge(entry: ManualKnowledgeEntry):
    """Add a single knowledge entry manually"""
    pool = await get_db_pool()
    if not pool:
        raise HTTPException(status_code=500, detail="Database not available")

    if len(entry.content) < 50:
        raise HTTPException(status_code=400, detail="Content too short (min 50 characters)")

    # Get supplier_id if supplier_name provided
    supplier_id = None
    if entry.supplier_name:
        async with pool.acquire() as conn:
            supplier = await conn.fetchrow(
                "SELECT id FROM suppliers WHERE LOWER(name) LIKE LOWER($1)",
                f"%{entry.supplier_name}%"
            )
            if supplier:
                supplier_id = supplier['id']

    doc_id, is_new = await store_knowledge(
        pool=pool,
        title=entry.title,
        content=entry.content,
        source_type="manual",
        source_name="Manual Entry",
        category=entry.category,
        subcategory=entry.subcategory,
        tags=entry.tags,
        supplier_id=supplier_id
    )

    return {
        "status": "success",
        "document_id": doc_id,
        "is_new": is_new,
        "chunks_created": len(chunk_text(entry.content))
    }


@router.post("/file")
async def upload_knowledge_file(
    file: UploadFile = File(...),
    category: str = Form("general"),
    subcategory: str = Form(None),
    tags: str = Form(None)  # Comma-separated
):
    """Upload a file (PDF, TXT, MD) to the knowledge base"""
    pool = await get_db_pool()
    if not pool:
        raise HTTPException(status_code=500, detail="Database not available")

    # Validate file type
    filename = file.filename.lower()
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''

    if ext not in ['pdf', 'txt', 'md']:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {ext}. Supported: pdf, txt, md"
        )

    # Read file content
    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 50MB)")

    # Extract text based on file type
    if ext == 'pdf':
        text = await extract_text_from_pdf(content)
    else:
        text = content.decode('utf-8', errors='ignore')

    if len(text) < 100:
        raise HTTPException(status_code=400, detail="File has insufficient text content")

    # Parse tags
    tag_list = [t.strip() for t in tags.split(',')] if tags else None

    # Store in knowledge base
    doc_id, is_new = await store_knowledge(
        pool=pool,
        title=file.filename,
        content=text,
        source_type="file_upload",
        source_name=file.filename,
        category=category,
        subcategory=subcategory,
        tags=tag_list
    )

    return {
        "status": "success",
        "document_id": doc_id,
        "is_new": is_new,
        "filename": file.filename,
        "text_length": len(text),
        "chunks_created": len(chunk_text(text))
    }


@router.post("/csv")
async def import_csv_knowledge(
    file: UploadFile = File(...),
    category: str = Form("plates")
):
    """
    Import knowledge from CSV file.

    Expected columns for plates: name, description, specifications, supplier, applications
    Expected columns for general: title, content, category, tags
    """
    pool = await get_db_pool()
    if not pool:
        raise HTTPException(status_code=500, detail="Database not available")

    content = await file.read()
    text = content.decode('utf-8', errors='ignore')

    # Parse CSV
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        raise HTTPException(status_code=400, detail="CSV file is empty")

    results = {"imported": 0, "skipped": 0, "errors": []}

    for i, row in enumerate(rows):
        try:
            if category == "plates":
                # Plate-specific import
                title = row.get('name') or row.get('plate_name') or row.get('title', f'Plate {i+1}')

                # Build content from available fields
                content_parts = []
                if row.get('description'):
                    content_parts.append(row['description'])
                if row.get('specifications'):
                    content_parts.append(f"Specifications: {row['specifications']}")
                if row.get('applications'):
                    content_parts.append(f"Applications: {row['applications']}")
                if row.get('processing'):
                    content_parts.append(f"Processing: {row['processing']}")
                if row.get('thickness'):
                    content_parts.append(f"Thickness: {row['thickness']}")
                if row.get('hardness'):
                    content_parts.append(f"Hardness: {row['hardness']} Shore A")

                knowledge_content = '\n'.join(content_parts) if content_parts else row.get('content', '')
                supplier_name = row.get('supplier')
                tags = [t.strip() for t in row.get('tags', '').split(',')] if row.get('tags') else None

            else:
                # General knowledge import
                title = row.get('title', f'Entry {i+1}')
                knowledge_content = row.get('content', '')
                supplier_name = row.get('supplier')
                tags = [t.strip() for t in row.get('tags', '').split(',')] if row.get('tags') else None

            if len(knowledge_content) < 50:
                results['skipped'] += 1
                continue

            # Get supplier_id
            supplier_id = None
            if supplier_name:
                async with pool.acquire() as conn:
                    supplier = await conn.fetchrow(
                        "SELECT id FROM suppliers WHERE LOWER(name) LIKE LOWER($1)",
                        f"%{supplier_name}%"
                    )
                    if supplier:
                        supplier_id = supplier['id']

            doc_id, is_new = await store_knowledge(
                pool=pool,
                title=title,
                content=knowledge_content,
                source_type="csv_import",
                source_name=file.filename,
                category=category,
                tags=tags,
                supplier_id=supplier_id
            )

            if is_new:
                results['imported'] += 1
            else:
                results['skipped'] += 1

        except Exception as e:
            results['errors'].append(f"Row {i+1}: {str(e)}")

    return {
        "status": "success",
        "filename": file.filename,
        "total_rows": len(rows),
        **results
    }


@router.post("/openai-file")
async def upload_to_openai_filesearch(
    file: UploadFile = File(...),
    purpose: str = Form("assistants")
):
    """
    Upload a file directly to OpenAI for use with file_search tool.
    This allows the FlexoBrain assistant to search through uploaded documents.
    """
    if not os.getenv("OPENAI_API_KEY"):
        raise HTTPException(status_code=500, detail="OpenAI API key not configured")

    # Validate file type - OpenAI supports various document types
    filename = file.filename.lower()
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''

    supported = ['pdf', 'txt', 'md', 'docx', 'doc', 'json', 'csv']
    if ext not in supported:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type for OpenAI: {ext}. Supported: {', '.join(supported)}"
        )

    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 50MB)")

    try:
        client = get_openai_client()

        # Upload file to OpenAI
        openai_file = await client.files.create(
            file=(file.filename, content),
            purpose=purpose
        )

        return {
            "status": "success",
            "openai_file_id": openai_file.id,
            "filename": file.filename,
            "bytes": openai_file.bytes,
            "purpose": openai_file.purpose,
            "message": "File uploaded to OpenAI. Add to a vector store to use with file_search."
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OpenAI upload failed: {str(e)}")


@router.post("/seed")
async def seed_industry_knowledge(request: SeedKnowledgeRequest, background_tasks: BackgroundTasks):
    """
    Seed the knowledge base with comprehensive flexographic printing knowledge.
    This provides a foundation of industry expertise for FlexoBrain.
    """
    pool = await get_db_pool()
    if not pool:
        raise HTTPException(status_code=500, detail="Database not available")

    knowledge_entries = []

    if request.include_core_knowledge:
        knowledge_entries.extend(CORE_FLEXO_KNOWLEDGE)

    if request.include_supplier_info:
        knowledge_entries.extend(SUPPLIER_KNOWLEDGE)

    if request.include_troubleshooting:
        knowledge_entries.extend(TROUBLESHOOTING_KNOWLEDGE)

    if request.include_best_practices:
        knowledge_entries.extend(BEST_PRACTICES_KNOWLEDGE)

    async def seed_all():
        results = {"added": 0, "skipped": 0}
        for entry in knowledge_entries:
            try:
                doc_id, is_new = await store_knowledge(
                    pool=pool,
                    title=entry['title'],
                    content=entry['content'],
                    source_type="seed",
                    source_name="FlexoBrain Industry Knowledge",
                    category=entry['category'],
                    subcategory=entry.get('subcategory'),
                    tags=entry.get('tags', [])
                )
                if is_new:
                    results['added'] += 1
                else:
                    results['skipped'] += 1
            except Exception as e:
                print(f"Error seeding {entry['title']}: {e}")
        return results

    # Run in background for large seed operations
    background_tasks.add_task(seed_all)

    return {
        "status": "seeding_started",
        "entries_queued": len(knowledge_entries),
        "message": "Knowledge seeding started in background"
    }


@router.get("/templates/csv")
async def get_csv_template(template_type: str = "plates"):
    """Get a CSV template for bulk import"""

    if template_type == "plates":
        template = """name,supplier,description,thickness,hardness,applications,processing,tags
nyloflex FTF 1.14,XSYS,Digital flat-top flexo plate for flexible packaging,1.14mm,68 Shore A,"flexible packaging,labels",solvent,"flat-top,digital,high-quality"
Cyrel EASY EFX 1.14,DuPont,Thermal flat-top digital plate,1.14mm,67 Shore A,"flexible packaging,labels,shrink",thermal,"flat-top,thermal,FAST"
"""
    else:
        template = """title,content,category,subcategory,tags
UV Exposure Basics,"Main exposure polymerizes image areas. Back exposure creates floor. Always run step tests.",processing,exposure,"UV,exposure,basics"
Dot Gain Control,"Control dot gain through proper exposure and anilox selection. Typical dot gain 15-25%.",troubleshooting,print_quality,"dot gain,TVI,quality"
"""

    return {
        "template_type": template_type,
        "csv_content": template,
        "instructions": "Download this template, fill in your data, and upload via /api/knowledge/load/csv"
    }


# ============================================================================
# SEED KNOWLEDGE DATA
# ============================================================================

CORE_FLEXO_KNOWLEDGE = [
    {
        "title": "Flexographic Printing Overview",
        "category": "best_practices",
        "subcategory": "fundamentals",
        "tags": ["introduction", "basics", "overview"],
        "content": """Flexographic printing (flexo) is a rotary relief printing process that uses flexible photopolymer printing plates. It is the dominant printing method for packaging, producing everything from flexible packaging and labels to corrugated boxes and folding cartons.

Key characteristics of flexo:
- Uses fluid inks (water-based, solvent-based, or UV-curable)
- Can print on virtually any substrate (film, paper, foil, corrugated)
- High-speed production (up to 600+ meters/minute)
- Low-to-medium setup costs compared to gravure
- Excellent for medium to long run lengths

The flexo press consists of:
1. Plate cylinder - carries the printing plate
2. Anilox roller - ceramic engraved roller that meters ink
3. Doctor blade - removes excess ink from anilox
4. Impression cylinder - provides pressure for ink transfer
5. Unwind/rewind systems for web handling"""
    },
    {
        "title": "Photopolymer Plate Technology",
        "category": "plates",
        "subcategory": "technology",
        "tags": ["photopolymer", "chemistry", "digital", "analog"],
        "content": """Modern flexographic plates are made from photopolymer materials - UV-reactive polymers that harden (crosslink) when exposed to ultraviolet light.

Digital Plates:
- Use a laser-ablatable mask layer (LAMS)
- Laser removes mask in image areas, exposing polymer
- Higher resolution, sharper dots
- Can hold 1-2% highlights and 98-99% shadows
- Examples: all modern plates (Cyrel, nyloflex, FLEXCEL NX)

Analog Plates:
- Use photographic film negatives
- Film placed over plate, UV exposure through clear areas
- Lower resolution but still suitable for many applications
- Being phased out in most markets

Key plate properties:
- Thickness: 0.76mm to 6.35mm+ depending on application
- Hardness: 55-80 Shore A durometer
- Surface type: Flat-top vs round-top dots
- Processing: Solvent, thermal, or water-wash compatible"""
    },
    {
        "title": "Flat-Top Dot Technology",
        "category": "plates",
        "subcategory": "surface_technology",
        "tags": ["flat-top", "FTF", "EASY", "NX", "surface"],
        "content": """Flat-top dot technology revolutionized flexographic print quality by providing a flatter, more consistent printing surface on each dot.

How it works:
- Traditional round-top dots have a curved surface that transfers ink unevenly
- Flat-top dots have a plateau-like surface that provides consistent ink laydown
- Created through special exposure techniques or plate construction

Major flat-top technologies:
1. XSYS nyloflex FTF - Uses unique polymer formulation
2. DuPont Cyrel EASY - Created using special exposure sequence
3. Miraclon FLEXCEL NX - Uses lamination technology with thermal imaging mask
4. Asahi CleanPrint - Optimized surface energy for clean release

Benefits of flat-top dots:
- More consistent ink density across the image
- Better highlight reproduction (can hold 1% dots)
- Reduced dot gain/TVI (tone value increase)
- Improved solid ink density
- More consistent print quality over longer runs
- Better for process color printing (CMYK)"""
    },
    {
        "title": "Plate Processing Methods",
        "category": "processing",
        "subcategory": "methods",
        "tags": ["solvent", "thermal", "water-wash", "FAST", "processing"],
        "content": """Three main processing methods exist for flexographic plates:

1. SOLVENT WASHOUT (Traditional):
- Uses chemical solvents to remove uncured polymer
- Solvents: Perchloroethylene (perc), hydrocarbon-based, or newer bio-solvents
- Processing time: 45-90 minutes total cycle
- Requires solvent recovery and distillation systems
- Most established method, widest plate compatibility
- Environmental considerations: VOC emissions, disposal

2. THERMAL PROCESSING (FAST):
- Uses heat and absorbent media (nonwoven fabric)
- No liquid chemicals or drying time
- Processing time: 15-25 minutes total
- Creates consistent floor thickness
- More environmentally friendly
- Higher initial equipment cost
- Examples: DuPont Cyrel FAST, XSYS ThermoFlexx

3. WATER-WASH PROCESSING:
- Uses water with small amount of detergent
- Most environmentally friendly option
- Requires water treatment/recycling
- Specialized plates required (Asahi AWP, Toyobo, DuPont NOW)
- Processing time: 30-45 minutes
- Growing in popularity due to sustainability focus

Choosing a method depends on: environmental regulations, plate requirements, throughput needs, capital budget, and substrate applications."""
    },
    {
        "title": "UV Exposure Fundamentals",
        "category": "processing",
        "subcategory": "exposure",
        "tags": ["UV", "exposure", "main", "back", "energy"],
        "content": """UV exposure is the critical step that polymerizes (cures) the flexographic plate. Understanding exposure is essential for consistent plate quality.

MAIN EXPOSURE (Front Exposure):
- Purpose: Polymerize image areas through the mask
- Creates the printing surface
- Energy range: 800-2000 mJ/cm² (varies by plate type)
- Controls: dot size, shoulder angle, resolution
- Under-exposure: soft dots, poor durability, dot loss
- Over-exposure: dot gain, bridging, loss of fine detail

BACK EXPOSURE (Floor Exposure):
- Purpose: Create the plate floor/base
- Controls relief depth and floor thickness
- Energy range: 100-400 mJ/cm² typically
- Performed before main exposure
- Determines minimum dot holding capability

UV SOURCES:
1. Fluorescent UVA tubes
   - Output: 15-25 mW/cm² when new
   - Degrade over time (need replacement at 1000-2000 hours)
   - Require warm-up time
   - Lower initial cost

2. LED UVA
   - Output: 30-80 mW/cm² (higher intensity)
   - Consistent output over 20,000+ hour life
   - Instant on/off, no warm-up
   - Lower operating costs
   - Higher initial investment

EXPOSURE CALCULATION:
Time (seconds) = Energy (mJ/cm²) ÷ Intensity (mW/cm²)

Example: 1000 mJ/cm² ÷ 20 mW/cm² = 50 seconds"""
    },
    {
        "title": "Plate Thickness Selection Guide",
        "category": "plates",
        "subcategory": "selection",
        "tags": ["thickness", "selection", "applications"],
        "content": """Selecting the correct plate thickness is crucial for print quality and plate life. Thickness affects impression latitude, dot reproduction, and substrate compatibility.

THIN PLATES (0.76mm - 1.14mm):
- Applications: Labels, flexible packaging, high-quality process work
- LPI capability: 133-200 LPI
- Benefits: Best resolution, finest dots, lowest dot gain
- Considerations: Requires precise impression settings, less forgiving
- Examples: Film labels, shrink sleeves, high-end flexible packaging

MEDIUM PLATES (1.70mm):
- Applications: General flexible packaging, folding carton, some labels
- LPI capability: 100-150 LPI
- Benefits: Good balance of quality and forgiveness
- Considerations: Most versatile thickness
- Examples: Food packaging, pharmaceutical labels, flexible pouches

MEDIUM-THICK PLATES (2.54mm - 2.84mm):
- Applications: Folding carton, light corrugated, paper bags
- LPI capability: 85-133 LPI
- Benefits: More impression latitude, cushioning effect
- Considerations: Not for finest detail work

THICK PLATES (3.18mm - 6.35mm):
- Applications: Corrugated postprint, heavy corrugated
- LPI capability: 65-100 LPI
- Benefits: Maximum cushion for uneven substrates
- Considerations: Requires proper mounting, lower resolution
- Examples: Corrugated boxes, kraft paper

Rule of thumb: Use the thinnest plate that will work for your substrate and quality requirements."""
    },
    {
        "title": "Anilox Roller Selection",
        "category": "equipment",
        "subcategory": "anilox",
        "tags": ["anilox", "line screen", "BCM", "cell volume"],
        "content": """The anilox roller is the heart of the flexo ink system, precisely metering ink to the printing plate.

KEY SPECIFICATIONS:

1. Line Screen (cells per inch/cm):
- Higher line screen = finer cells = less ink, better detail
- Lower line screen = coarser cells = more ink, better coverage
- Typical range: 200-1200 cells per inch

2. Cell Volume (BCM or cm³/m²):
- Determines ink laydown amount
- BCM = Billion Cubic Microns per square inch
- Higher volume = more ink = denser solids
- Lower volume = less ink = cleaner highlights

SELECTION GUIDELINES:

Process Work (CMYK):
- Line screen: 800-1200 cpi
- Volume: 1.5-3.5 BCM
- Higher screen for better tones

Solid/Spot Colors:
- Line screen: 300-500 cpi
- Volume: 4-8 BCM
- Lower screen for better density

White/Opaque Inks:
- Line screen: 180-360 cpi
- Volume: 8-14 BCM
- Needs high volume for opacity

THE 5:1 TO 7:1 RULE:
Anilox line screen should be 5-7x the plate LPI
Example: 133 LPI plate → 665-931 cpi anilox

CELL GEOMETRY:
- 60° Hexagonal: Most common, good release
- 30° Channel: Better for high-viscosity inks
- Open/elongated cells: Better for metallic inks

MAINTENANCE:
- Regular cleaning essential (plugged cells reduce volume)
- Inspect under microscope
- Track effective volume over time
- Replace when worn (typically 3-5 years for ceramic)"""
    },
    {
        "title": "Shore A Hardness and Applications",
        "category": "plates",
        "subcategory": "hardness",
        "tags": ["hardness", "durometer", "Shore A", "selection"],
        "content": """Plate hardness (measured in Shore A durometer) significantly affects print quality, dot reproduction, and plate durability.

SOFT PLATES (55-62 Shore A):
- Better ink transfer on rough substrates
- More forgiving of impression variation
- Higher dot gain
- Applications: Corrugated, textured papers
- Trade-off: Less fine detail capability

MEDIUM PLATES (63-68 Shore A):
- Balanced performance
- Good for general packaging
- Moderate dot gain
- Applications: Flexible packaging, labels, folding carton
- Most common hardness range

HARD PLATES (69-75 Shore A):
- Better fine detail reproduction
- Lower dot gain
- Less forgiving of impression variation
- Applications: High-quality process work, labels
- Requires precise press setup

VERY HARD PLATES (76-80+ Shore A):
- Excellent highlight holding
- Minimal dot gain
- Highest resolution capability
- Applications: Security printing, fine labels
- Requires expert press operation

HARDNESS SELECTION FACTORS:

1. Substrate roughness:
   - Rough substrate → softer plate
   - Smooth substrate → harder plate

2. Print quality requirements:
   - Fine detail/process → harder plate
   - Solids/simple graphics → softer plate

3. Press condition:
   - Worn press → softer plate (more forgiving)
   - New press → can use harder plates

4. Run length:
   - Longer runs → harder plate (better durability)
   - Short runs → hardness less critical"""
    }
]

SUPPLIER_KNOWLEDGE = [
    {
        "title": "XSYS (formerly Flint Group Flexographic)",
        "category": "plates",
        "subcategory": "suppliers",
        "tags": ["XSYS", "nyloflex", "supplier"],
        "content": """XSYS (formerly Flint Group Flexographic Products) is a leading global supplier of flexographic plates and equipment.

KEY PRODUCT LINES:

nyloflex Plates:
- nyloflex FTF: Flat-top technology for premium quality
- nyloflex FAH/FAC: High durometer for corrugated
- nyloflex ACE: Thermal-processable (ThermoFlexx compatible)
- nyloflex FTV: Versatile general-purpose digital plate
- nyloflex Sprint: Fast processing solvent plate

Processing Types:
- Solvent: Most nyloflex varieties
- Thermal: ACE series with ThermoFlexx processor

Thickness Range: 0.76mm to 6.35mm

Key Differentiators:
- Strong in corrugated and flexible packaging
- FTF flat-top technology for high-quality printing
- ThermoFlexx thermal processing system
- Global support network

Best For:
- High-quality flexible packaging
- Corrugated postprint
- Printers seeking German engineering quality"""
    },
    {
        "title": "DuPont Cyrel",
        "category": "plates",
        "subcategory": "suppliers",
        "tags": ["DuPont", "Cyrel", "FAST", "EASY", "supplier"],
        "content": """DuPont Cyrel is one of the most recognized names in flexographic plates, pioneering many technologies used today.

KEY PRODUCT LINES:

Cyrel Plates:
- Cyrel EASY: Revolutionary flat-top dot technology
- Cyrel DFH/DFR: High-performance corrugated plates
- Cyrel NOW: Water-wash eco-friendly plates
- Cyrel DSE: Digital solvent plates

Processing Types:
- Cyrel FAST: Thermal processing (industry standard)
- Solvent: Traditional processing
- Water-wash: NOW series

Unique Technologies:
- EASY (Enhanced Appearance Substantially Yours): Flat-top dots through special exposure
- FAST: First commercially successful thermal processing
- DigiFlow: Advanced screening for smooth vignettes

Thickness Range: 0.76mm to 7.00mm

Key Differentiators:
- Industry pioneer with decades of experience
- Strong R&D and innovation
- Extensive technical support
- Global availability

Best For:
- Premium quality flexible packaging
- Printers wanting established technology
- Those invested in FAST thermal processing"""
    },
    {
        "title": "Miraclon (FLEXCEL NX)",
        "category": "plates",
        "subcategory": "suppliers",
        "tags": ["Miraclon", "FLEXCEL", "NX", "Kodak", "supplier"],
        "content": """Miraclon (formerly Kodak Flexographic Packaging Division) offers the premium FLEXCEL NX System known for exceptional print quality.

KEY PRODUCT LINE:

FLEXCEL NX System:
- Uses unique Thermal Imaging Layer (TIL)
- Lamination-based imaging process
- Creates ultra-precise flat-top dots
- Produces "DigiCaps" micro-textured surface

FLEXCEL NX Plates:
- NX Ultra: Maximum quality and durability
- NX Print: Balance of quality and versatility
- Various thicknesses for all applications

Processing:
- Solvent washout (standard nyloflex-type solvents)
- Not thermal compatible (proprietary lamination process)

Unique Technology:
- TIL (Thermal Imaging Layer): Separate imaging layer laminated to plate
- Digital Flexo Dot (DFD): Flat-top with controlled shoulders
- DigiCaps: Micro-texture for optimal ink release
- Finest highlight reproduction in industry (0.4% dots possible)

Key Differentiators:
- Highest quality in the market
- Premium pricing reflects premium results
- Proprietary technology (requires FLEXCEL NX imager)
- Exceptional consistency

Best For:
- Brand owners demanding best quality
- High-value flexible packaging
- Process color work with extended gamut
- Printers positioning as quality leaders"""
    },
    {
        "title": "Asahi Photoproducts",
        "category": "plates",
        "subcategory": "suppliers",
        "tags": ["Asahi", "AWP", "water-wash", "CleanPrint", "supplier"],
        "content": """Asahi Photoproducts is a Japanese-owned company leading in water-wash plate technology and sustainability.

KEY PRODUCT LINES:

AWP (Asahi Water-wash Plates):
- AWP-DEW: Digital water-wash for flexible packaging
- AWP-DEF: Digital water-wash for labels
- AWP CleanPrint: Optimized surface for clean ink release
- Various thicknesses and hardnesses

Processing:
- Water-wash only (no solvent required)
- Uses water with mild detergent
- Most environmentally friendly option

CleanPrint Technology:
- Engineered surface texture
- Improved ink release
- Reduced need for cleaning during press runs
- Consistent quality over long runs

Thickness Range: 0.76mm to 7.00mm

Key Differentiators:
- Leader in water-wash technology
- Strong sustainability story
- CleanPrint surface technology
- Good for printers with environmental mandates

Best For:
- Environmentally conscious printers
- Regions with strict VOC regulations
- Food packaging (no solvent residue concerns)
- Printers wanting to eliminate solvents"""
    }
]

TROUBLESHOOTING_KNOWLEDGE = [
    {
        "title": "Dot Gain / TVI Troubleshooting",
        "category": "troubleshooting",
        "subcategory": "print_quality",
        "tags": ["dot gain", "TVI", "quality", "troubleshooting"],
        "content": """Dot gain (Tone Value Increase/TVI) is the increase in dot size from plate to print. Some dot gain is normal in flexo (15-25%), but excessive gain causes problems.

SYMPTOMS:
- Midtones appear darker than intended
- Highlights fill in
- Loss of detail in shadows
- Colors appear oversaturated

CAUSES AND SOLUTIONS:

1. Excessive Impression Pressure:
- Cause: Too much squeeze between plate and substrate
- Solution: Reduce impression to "kiss" contact
- Test: Print solid bar, check for even density

2. Plate Issues:
- Over-exposed plates have larger dots
- Soft plates gain more than hard plates
- Solution: Check exposure, consider harder plate

3. Anilox Issues:
- Too much ink volume
- Plugged cells causing uneven ink
- Solution: Use lower volume anilox, clean cells

4. Ink Issues:
- Ink viscosity too low (thin ink spreads)
- Wrong ink/substrate combination
- Solution: Adjust viscosity, check ink compatibility

5. Mounting Issues:
- Improper tape/sleeve cushion
- Wrong tape thickness
- Solution: Review mounting tape selection

6. Environmental Factors:
- High humidity affects ink drying
- Temperature affects ink viscosity
- Solution: Control press room conditions

MEASUREMENT:
- Use densitometer with dot gain function
- Measure 50% patch typically
- Target: 15-25% gain for standard flexo"""
    },
    {
        "title": "Dirty Printing / Scumming",
        "category": "troubleshooting",
        "subcategory": "print_quality",
        "tags": ["dirty printing", "scumming", "haze", "troubleshooting"],
        "content": """Dirty printing (also called scumming or haze) occurs when ink appears in non-image areas.

SYMPTOMS:
- Ink in clear/non-printing areas
- Hazy appearance in reverses
- Ink buildup on plate floor
- Progressively worsens during run

CAUSES AND SOLUTIONS:

1. Over-Impression:
- Plate floor contacts substrate
- Solution: Reduce impression pressure

2. Plate Floor Issues:
- Floor too thin from insufficient back exposure
- Debris embedded in floor
- Solution: Increase back exposure, ensure clean platemaking

3. Ink Issues:
- Ink too thin/low viscosity
- Incompatible ink chemistry
- Solution: Increase viscosity, check ink compatibility

4. Anilox Issues:
- Anilox volume too high
- Ink being pushed to non-image areas
- Solution: Use lower volume anilox

5. Doctor Blade Problems:
- Blade not removing excess ink
- Blade damaged or worn
- Solution: Check/replace doctor blade

6. Static Electricity:
- Attracts ink to non-image areas
- Common with film substrates
- Solution: Install static eliminators

7. Plate Swelling:
- Some inks cause plate to swell
- Creates contact in non-image areas
- Solution: Check ink/plate compatibility

PREVENTION:
- Proper exposure times (especially back exposure)
- Correct impression settings
- Appropriate anilox selection
- Regular doctor blade inspection"""
    },
    {
        "title": "Plate Wear and Durability Issues",
        "category": "troubleshooting",
        "subcategory": "plates",
        "tags": ["plate wear", "durability", "degradation", "troubleshooting"],
        "content": """Plate wear reduces print quality over time. Understanding wear mechanisms helps maximize plate life.

SYMPTOMS:
- Highlight dots wearing away
- Fine lines becoming thinner
- Density loss in solids
- Edges becoming rounded

CAUSES AND SOLUTIONS:

1. Abrasive Substrates:
- Paper fibers, foil, rough materials
- Solution: Use harder plate, reduce impression

2. Excessive Impression:
- Over-pressure accelerates wear
- Solution: Set minimum impression, use softer tape

3. Ink Chemistry:
- Some inks attack plate polymer
- UV inks can be more aggressive
- Solution: Verify ink/plate compatibility

4. Anilox Contact:
- Worn ceramic can abrade plate
- Solution: Inspect anilox, maintain doctor blade

5. Plate Quality:
- Poor exposure creates weak dot structure
- Solution: Optimize exposure parameters

6. Static Issues:
- Can attract debris that abrades plate
- Solution: Use static elimination

MAXIMIZING PLATE LIFE:

1. Proper Storage:
- Store flat, away from UV light
- Controlled temperature (15-25°C)
- Away from ozone sources

2. Handling:
- Clean hands/gloves
- Proper mounting/demounting
- Avoid bending or creasing

3. Cleaning:
- Use compatible plate wash
- Gentle cleaning, no abrasives
- Dry properly before storage

4. Run Practices:
- Minimum impression
- Correct ink viscosity
- Regular anilox cleaning

EXPECTED PLATE LIFE:
- Film substrates: 1-2 million impressions
- Paper substrates: 500K-1 million impressions
- Abrasive substrates: 100K-500K impressions"""
    },
    {
        "title": "Exposure Problems Diagnosis",
        "category": "troubleshooting",
        "subcategory": "platemaking",
        "tags": ["exposure", "under-exposure", "over-exposure", "troubleshooting"],
        "content": """Exposure problems affect every aspect of plate performance. Learn to diagnose and correct exposure issues.

UNDER-EXPOSURE SYMPTOMS:
Main exposure:
- Soft, mushy dots
- Dots wash away during processing
- Poor plate durability
- Fuzzy dot edges
- Loss of fine positive elements

Back exposure:
- Excessive relief depth
- Thin floor
- Plate feels flexible
- Highlights wash away

OVER-EXPOSURE SYMPTOMS:
Main exposure:
- Dot gain on plate
- Loss of fine reverses
- Dots bridging together
- Halftone range compressed
- Light-trap in reverses

Back exposure:
- Relief too shallow
- Floor too thick
- Cannot achieve proper relief
- Difficulty washing out

DIAGNOSIS STEPS:

1. Run Step Test (UGRA/FOGRA wedge):
- Measure minimum reproducible dot
- Measure solid step reproduction
- Compare to manufacturer specs

2. Check UV Intensity:
- Use radiometer
- Compare to expected values
- Check uniformity across exposure area

3. Inspect Processing:
- Ensure complete washout
- Check for residue
- Verify drying/finishing

CORRECTION:

Under-exposed plates:
- Increase exposure time
- Check lamp output (may need replacement)
- Verify vacuum drawdown

Over-exposed plates:
- Reduce exposure time
- Check for light leaks
- Verify mask/LAMS quality

EXPOSURE TESTING SCHEDULE:
- Daily: Visual check of test patches
- Weekly: Full step test measurement
- Monthly: UV intensity measurement
- Quarterly: Full audit of all parameters"""
    }
]

BEST_PRACTICES_KNOWLEDGE = [
    {
        "title": "Plate Mounting Best Practices",
        "category": "best_practices",
        "subcategory": "mounting",
        "tags": ["mounting", "tape", "registration", "best practices"],
        "content": """Proper plate mounting is critical for print quality and registration. Follow these best practices.

MOUNTING TAPE SELECTION:

Foam Tape Types:
- Soft (20-25 Shore A): Maximum cushion, corrugated
- Medium (30-35 Shore A): General packaging
- Hard (40-50 Shore A): High-quality process work

Tape Thickness:
- 0.38mm (15 mil): Thin plates, tight tolerances
- 0.50mm (20 mil): Most common, general purpose
- 0.75mm (30 mil): Extra cushion for rough substrates

MOUNTING PROCEDURE:

1. Prepare Mounting Surface:
- Clean sleeve/cylinder thoroughly
- Remove any old adhesive residue
- Ensure surface is dry

2. Apply Tape:
- Cut tape precisely
- Apply to plate back, not cylinder (for foam tapes)
- Use roller to eliminate air bubbles
- Ensure complete adhesion

3. Position Plate:
- Use registration marks
- Pin registration preferred for multi-color
- Video registration systems for tight tolerance work

4. Final Check:
- Verify no air bubbles under plate
- Check edge adhesion
- Confirm registration to other colors

COMMON MOUNTING PROBLEMS:

Air Bubbles:
- Cause print defects
- Solution: Use roller, work from center out

Poor Adhesion:
- Plate lifts during press run
- Solution: Clean surfaces, use fresh tape

Registration Errors:
- Colors don't align
- Solution: Use pin register, check marks

STICKY-BACK (Self-Adhesive) vs TAPE:

Sticky-back plates:
- Pre-applied adhesive
- Quick mounting
- Less control over cushion
- Good for short runs

Separate tape mounting:
- Choose cushion level
- More labor intensive
- Better for long runs
- Preferred for quality work"""
    },
    {
        "title": "Ink Management Best Practices",
        "category": "best_practices",
        "subcategory": "ink",
        "tags": ["ink", "viscosity", "management", "best practices"],
        "content": """Proper ink management ensures consistent print quality and reduces waste.

VISCOSITY CONTROL:

Why It Matters:
- Too thin: Dirty printing, poor coverage
- Too thick: Incomplete transfer, bridging

Measurement:
- Use Zahn or efflux cup
- Measure at consistent temperature
- Document target values by color

Adjustment:
- Add solvent/water to thin
- Add fresh ink to thicken
- Make small adjustments, re-measure

Typical Targets:
- Water-based: 20-30 sec (Zahn #2)
- Solvent: 18-25 sec (Zahn #2)
- UV: Specified by manufacturer

TEMPERATURE CONTROL:

Effects:
- Warm ink is thinner
- Cold ink is thicker
- 5°C change = significant viscosity change

Best Practices:
- Maintain press room at 20-25°C
- Pre-condition ink before use
- Use ink temperature control systems

COLOR MANAGEMENT:

Batch Consistency:
- Check new batches against standard
- Use spectrophotometer for measurement
- Document Delta E values

Press Matching:
- Create press profiles
- Adjust formulations for specific presses
- Document ink drawdowns

INK STORAGE:

Water-based:
- Store away from freezing
- Keep containers sealed (skinning)
- Use FIFO (first in, first out)
- Typical shelf life: 6-12 months

Solvent-based:
- Well-ventilated storage
- Away from ignition sources
- Containers tightly sealed
- Follow safety protocols

UV Inks:
- Store away from light
- Temperature controlled
- Check expiration dates
- Handle with care (sensitizers)"""
    },
    {
        "title": "Quality Control in Platemaking",
        "category": "best_practices",
        "subcategory": "quality",
        "tags": ["quality control", "QC", "testing", "best practices"],
        "content": """Implementing quality control in platemaking ensures consistent, high-quality plates.

INCOMING MATERIAL CHECK:

Raw Plates:
- Check batch/lot numbers
- Verify correct type ordered
- Inspect for physical damage
- Check expiration dates

Processing Chemicals:
- Verify correct chemistry
- Check concentration/dilution
- Monitor activity levels

PROCESS CONTROL:

Exposure Monitoring:
- Daily intensity check
- Weekly step test
- Monthly uniformity check
- Track lamp hours

Processing Monitoring:
- Solution temperature
- Brush condition (solvent systems)
- Filter condition
- Dryer temperature

PLATE INSPECTION:

Visual Checks:
- Complete washout
- No debris in floor
- Dot formation
- Edge quality

Measurement:
- Floor thickness
- Relief depth
- Dot reproduction (step test)
- Registration marks

Tools Needed:
- Depth gauge/micrometer
- Magnifying glass/loupe (10-20x)
- Digital microscope (optional but valuable)
- Densitometer

DOCUMENTATION:

Track for Each Plate:
- Plate type and batch
- Exposure times (main/back)
- Processing parameters
- Inspection results
- Operator initials

Benefits of Documentation:
- Traceability for problems
- Process improvement data
- Training reference
- Customer requirements

CALIBRATION:

Equipment to Calibrate:
- UV radiometer
- Depth gauge
- Thermometers
- Densitometer

Frequency:
- Radiometer: Annually
- Mechanical gauges: Semi-annually
- Process monitors: Per manufacturer

Keep calibration records for audits."""
    },
    {
        "title": "Press Setup for Optimal Print Quality",
        "category": "best_practices",
        "subcategory": "press",
        "tags": ["press setup", "impression", "registration", "best practices"],
        "content": """Proper press setup is essential for achieving optimal print quality in flexography.

IMPRESSION SETTING:

The "Kiss" Impression:
- Minimum pressure that transfers ink
- Plate dots just contact substrate
- No visible squeeze-out

Setting Procedure:
1. Start with zero impression (gap)
2. Slowly increase until ink transfers
3. Stop at first sign of complete coverage
4. Fine-tune for quality

Common Mistakes:
- Over-impression (causes dot gain, wear)
- Under-impression (missed dots, inconsistent)

ANILOX-TO-PLATE SETTINGS:

Principles:
- Metered amount of ink to plate
- Consistent across web
- No excess causing dirty printing

Setting:
1. Start with light contact
2. Adjust for even ink coverage
3. Avoid heavy pressure (damages both)

REGISTRATION:

Mechanical Registration:
- Check gear alignment
- Verify cylinder runout
- Confirm web tension

In-Register Running:
- Use registration marks
- Camera systems for monitoring
- Make small adjustments

Troubleshooting:
- Gear-bound: Check gear mesh
- Random: Check web tension
- Progressive: Check temp/humidity

WEB TENSION:

Importance:
- Affects registration
- Affects impression
- Affects material stretch

Guidelines:
- Film: Lower tension (stretch risk)
- Paper: Higher tension acceptable
- Consistent throughout press

PRINT SPEED:

Considerations:
- Ink drying capability
- Material handling
- Quality requirements

Speed vs Quality:
- Slower = more control
- Faster = more efficient
- Find optimal balance

STARTUP CHECKLIST:

1. Verify job specifications
2. Check plate mounting
3. Set anilox engagement
4. Set impression
5. Check ink supply and viscosity
6. Verify web path and tension
7. Run waste to color
8. Fine-tune registration
9. Approve first article
10. Document settings"""
    }
]


@router.get("/health")
async def health_check():
    """Health check for knowledge loader"""
    pool = await get_db_pool()

    return {
        "status": "healthy",
        "database_connected": pool is not None,
        "openai_configured": bool(os.getenv("OPENAI_API_KEY"))
    }
